import csv
import os
import openpyxl
from openpyxl.utils import get_column_letter


class FileOutputHelper:

    @staticmethod
    def append_to_csv(response_info):
        """Add data to local csv"""
        header = [
            "query",
            "context_text",
            "context_time_ms",
            "response_text",
            "response_time_ms",
            "db_time_ms",
            "similarity_results",
        ]
        row = [
            response_info["query"],
            response_info["context_text"],
            response_info["search_time"],
            response_info["response_text"],
            response_info["response_time"],
            response_info["setup_db_time"],
            response_info["retrieved_items"]  # No longer have Similarity Score as retriever interaction changed
        ]
        logger_file_path = response_info["logger_file_path"]

        try:
            os.makedirs(os.path.dirname(logger_file_path), exist_ok=True)
            file_exists = os.path.isfile(logger_file_path)
            with open(logger_file_path, mode="a", newline="", encoding="utf-8") as file:
                writer = csv.writer(file)
                if not file_exists:  # If file does not exist, write the header
                    writer.writerow(header)
                writer.writerow(row)
        except IOError as e:
            print(f"An error occurred while writing to the CSV file: {e}")


    @staticmethod
    def save_retrieved_to_logger(doc_name, retriever_type_lst, record, extension_name=".xlsx"):
        # create folder
        folder_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__))) + "/logger/retrieved/"
        retrievers_num = len(retriever_type_lst)
        file_name = doc_name + "_" + str(retrievers_num) + "retriever" + extension_name
        os.makedirs(folder_path, exist_ok=True)

        # create file object & config format
        wb = openpyxl.Workbook()
        ws = wb.active
        headers = ["Query_ID", "Query", "Chunks"]
        ws.append(headers)
        current_row = 2
        column_width = 25
        row_height = 200
        for col in range(1, len(headers) + 10):
            ws.column_dimensions[get_column_letter(col)].width = column_width

        # input record data
        for idx, retrieved_dict in record.items():
            max_chunks = max(len(sub_list) for sub_list in retrieved_dict["retrieved_chunks"].values())

            # fill metadata
            ws.cell(row=current_row, column=1, value=idx)
            ws.cell(row=current_row, column=2, value=retrieved_dict["question"])
            for i, rt in enumerate(retriever_type_lst):
                ws.cell(row=current_row + i, column=3, value=rt)
                ws.row_dimensions[current_row + i].height = row_height
                # fill chunks content
                cur_retriever_chunks = retrieved_dict["retrieved_chunks"][rt]
                for j in range(max_chunks):
                    if j < len(cur_retriever_chunks):
                        ws.cell(row=current_row + i, column=4 + j, value=cur_retriever_chunks[j])

            # format this query in Excel
            if retrievers_num > 1:
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + retrievers_num - 1, end_column=1)
                ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + retrievers_num - 1, end_column=2)
            ws.row_dimensions[current_row].height = row_height

            current_row += retrievers_num

        # save file
        file_abs_path = folder_path + file_name
        wb.save(file_abs_path)
        print("File '{}' saved".format(file_name))


    @staticmethod
    def move_forwards_same_items(lst1, lst2):
        common_elements = set(lst1) & set(lst2)

        def custom_sort(item):
            return item not in common_elements, item

        sorted_list1 = sorted(lst1, key=custom_sort)
        sorted_list2 = sorted(lst2, key=custom_sort)

        return sorted_list1, sorted_list2
