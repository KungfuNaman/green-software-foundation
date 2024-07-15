import sys
import csv
import os
import pymongo
import itertools
import pandas as pd
import shutil
import openpyxl
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
from datetime import datetime

load_dotenv()
print("hello")


def append_to_csv(
    query, context, search_time, response, response_time, db_time, similarity_results,logger_file_path
):
    """add data to local csv and mongo cloud"""
    header = [
        "query",
        "context_text",
        "context_time_ms",
        "response_text",
        "response_time_ms",
        "db_time_ms",
        "similarity_results",
    ]
    row = [
        query,
        context,
        search_time,
        response,
        response_time,
        db_time,
        similarity_results,
    ]

    try:
        # Ensure the directory exists
        os.makedirs(os.path.dirname(logger_file_path), exist_ok=True)

        file_exists = os.path.isfile(logger_file_path)

        with open(logger_file_path, mode="a", newline="", encoding="utf-8") as file:
            writer = csv.writer(file)
            if not file_exists:  # If file does not exist, write the header
                writer.writerow(header)
            writer.writerow(row)
    except IOError as e:
        print(f"An error occurred while writing to the CSV file: {e}")
        # add_to_mongo(query,context,search_time,response,response_time,db_time,similarity_results)


def add_to_mongo(query,context,search_time,response,response_time,db_time,similarity_results):
    try:
        collection = get_mongo_collection()
        document = {
            "query": query,
            "context_text": context,
            "context_time_ms": search_time,
            "response_text": response,
            "response_time_ms": response_time,
            "db_time_ms": db_time,
            "similarity_results": serialize_document(similarity_results),
        }

        try:
            # Insert the document into the MongoDB collection
            collection.insert_one(document)
            print("Document inserted successfully.")
        except pymongo.errors.PyMongoError as e:
            print(f"An error occurred while inserting the document into MongoDB: {e}")
    except Exception as e:
        print(f"An error occurred while interacting with MongoDB: {e}")

MONGO_STRING = os.getenv("MONGO_STRING")
MONGO_COLLECTION = os.getenv("MONGO_COLLECTION")
MONGO_DATABASE = os.getenv("MONGO_DATABASE")


def get_mongo_collection():
    """Set up and return the MongoDB collection"""
    client = pymongo.MongoClient(MONGO_STRING)
    db = client[MONGO_DATABASE]
    collection = db[MONGO_COLLECTION]
    return collection


def serialize_document(documents):
    obj = {}
    obj["meta_data"] = [doc.metadata.get("id", None) for doc, _score in documents]
    return obj


def convert_data_to_csv():
    collection = get_mongo_collection()
    # Query all documents in the collection
    documents = list(collection.find())

    # Create a DataFrame from the documents
    df = pd.DataFrame(documents)

    # Export DataFrame to CSV
    df.to_csv("gsf_export.csv", index=False)


def import_csv_to_mongo(csv_file_path):
    # Read the CSV file into a DataFrame
    df = pd.read_csv(csv_file_path)

    # Convert the DataFrame to a list of dictionaries
    data = df.to_dict(orient="records")

    # Get the MongoDB collection
    collection = get_mongo_collection()

    # Insert the list of dictionaries into the MongoDB collection
    if data:
        collection.insert_many(data)
        print(f"Inserted {len(data)} documents into the collection.")
    else:
        print("No data found in the CSV file.")


def save_retrieved_to_logger(doc_name, ds_type, record, extension_name=".xlsx"):
    # create folder
    folder_path = "Rag/logger/retrieved/"
    if record[0]["new_prediction"]:
        file_name = doc_name + "_" + ds_type + "_" + record[0]["retriever_type"] + "_NC" + extension_name
    else:
        file_name = doc_name + "_" + ds_type + "_" + record[0]["retriever_type"] + extension_name
    os.makedirs(folder_path, exist_ok=True)

    # create file object
    wb = openpyxl.Workbook()
    ws = wb.active
    headers = ["query_id", "query", "prediction", "ground_truth", "chunk"]
    ws.append(headers)
    current_row = 2
    column_width = 25
    row_height = 200
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = column_width

    # input record data
    for idx, retrieved_dict in record.items():
        ws.cell(row=current_row, column=1, value=idx)
        ws.cell(row=current_row, column=2, value=retrieved_dict["question"])
        ws.cell(row=current_row, column=3, value=retrieved_dict["prediction"])
        ws.cell(row=current_row, column=4, value=retrieved_dict["truth"])
        chroma_chunks = retrieved_dict["chroma_chunks"]
        llm_chunks = retrieved_dict["llm_chunks"]
        chroma_chunks, llm_chunks = move_forwards_same_items(chroma_chunks, llm_chunks)

        # adapts to different lengths
        max_chunks = max(len(chroma_chunks), len(llm_chunks))
        for i in range(max_chunks):
            if i < len(chroma_chunks):
                ws.cell(row=current_row, column=5 + i, value=chroma_chunks[i])
            if i < len(llm_chunks):
                ws.cell(row=current_row + 1, column=5 + i, value=llm_chunks[i])

        # format excel
        ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row + 1, end_column=1)
        ws.merge_cells(start_row=current_row, start_column=2, end_row=current_row + 1, end_column=2)
        ws.merge_cells(start_row=current_row, start_column=3, end_row=current_row + 1, end_column=3)
        ws.merge_cells(start_row=current_row, start_column=4, end_row=current_row + 1, end_column=4)
        ws.row_dimensions[current_row].height = row_height
        ws.row_dimensions[current_row + 1].height = row_height

        current_row += 2

    # save file
    file_abs_path = os.path.dirname(os.path.abspath(__file__)) + "/retrieved/" + file_name
    wb.save(file_abs_path)
    print("File '{}' saved".format(file_name))


def move_forwards_same_items(lst1, lst2):
    common_elements = set(lst1) & set(lst2)

    def custom_sort(item):
        return item not in common_elements, item

    sorted_list1 = sorted(lst1, key=custom_sort)
    sorted_list2 = sorted(lst2, key=custom_sort)

    return sorted_list1, sorted_list2



# import_csv_to_mongo("Rag/logger/llmResponse.csv")
# convert_data_to_csv()
